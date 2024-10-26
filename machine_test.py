import os
import openpyxl # library to read/write excel file

# file name for storing user data
file_name="data_of_users.xlsx"

# function to create a new excel file with defined file name if it doesn't exist
def create_file():
    if not os.path.exists(file_name):
        workbook=openpyxl.Workbook()
        sheet=workbook.active
        sheet.append(["Name", "Email", "Phone Number"]) # headers for the Excel file
        workbook.save(file_name) # saving the file

# function to add a user to the excel file
def add_user():
    name=input("Enter Name: ")
    email=input("Enter Email: ")
    phone=input("Enter Phone Number: ")
    workbook=openpyxl.load_workbook(file_name) # opening the excel file
    sheet=workbook.active
    sheet.append([name, email, phone]) # appending the user data to the file
    workbook.save(file_name)
    print("User added successfully!\n")

# function to display users from the excel file
def display_users():
    workbook=openpyxl.load_workbook(file_name)
    sheet=workbook.active
    print("\nStored Users:")
    # below for loop reads and display each row of data from the excel file
    for row in sheet.iter_rows(values_only=True):
        print(f"Name: {row[0]}, Email: {row[1]}, Phone Number: {row[2]}")

# main function
def main():
    create_file()
    while True:
        print("Make a choice to proceed:")
        print("Press '1' to add user")
        print("Press '2' to display users")
        print("Press '3' to exit")

        choice=input("Press any key to continue: ")     
        if choice=='1':
            add_user()
        elif choice=='2':
            display_users()
        elif choice=='3':
            print("Program terminated.")
            break
        else:
            print("Invalid choice, please try again.\n")

if __name__ == "__main__":
    main()
